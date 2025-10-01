import pandas as pd
import io
from datetime import datetime
from typing import Dict
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.lib.enums import TA_CENTER, TA_LEFT
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows


class ReportGenerator:
    """Generates exportable reports in PDF and Excel formats"""
    
    def __init__(self, full_data: pd.DataFrame, categorized_data: Dict[str, pd.DataFrame], analyzer):
        self.full_data = full_data
        self.categorized_data = categorized_data
        self.analyzer = analyzer
        self.styles = getSampleStyleSheet()
        self._setup_custom_styles()
    
    def _setup_custom_styles(self):
        """Setup custom styles for PDF generation"""
        # Title style
        self.styles.add(ParagraphStyle(
            name='CustomTitle',
            parent=self.styles['Heading1'],
            fontSize=24,
            textColor=colors.HexColor('#1f77b4'),
            spaceAfter=30,
            alignment=TA_CENTER
        ))
        
        # Section header style
        self.styles.add(ParagraphStyle(
            name='SectionHeader',
            parent=self.styles['Heading2'],
            fontSize=16,
            textColor=colors.HexColor('#2c3e50'),
            spaceAfter=12,
            spaceBefore=12,
            borderWidth=2,
            borderColor=colors.HexColor('#1f77b4'),
            borderPadding=5
        ))
        
        # Insight style
        self.styles.add(ParagraphStyle(
            name='Insight',
            parent=self.styles['Normal'],
            fontSize=10,
            textColor=colors.HexColor('#34495e'),
            leftIndent=20,
            spaceAfter=6
        ))
    
    def generate_pdf_report(self) -> bytes:
        """Generate a comprehensive PDF report"""
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=letter, rightMargin=72, leftMargin=72,
                               topMargin=72, bottomMargin=18)
        
        story = []
        
        # Title
        title = Paragraph("Qualcomm Financial Analysis Report", self.styles['CustomTitle'])
        story.append(title)
        
        # Report date
        date_text = Paragraph(
            f"<i>Generated on: {datetime.now().strftime('%B %d, %Y')}</i>",
            self.styles['Normal']
        )
        story.append(date_text)
        story.append(Spacer(1, 0.3*inch))
        
        # Executive Summary
        story.append(Paragraph("Executive Summary", self.styles['SectionHeader']))
        insights = self.analyzer.generate_executive_summary()
        for insight in insights:
            story.append(Paragraph(f"• {insight}", self.styles['Insight']))
        story.append(Spacer(1, 0.3*inch))
        
        # Financial Health Score
        health_score = self.analyzer.calculate_financial_health_score()
        if health_score and 'Overall' in health_score:
            story.append(Paragraph("Financial Health Assessment", self.styles['SectionHeader']))
            story.append(Paragraph(
                f"<b>Overall Score: {health_score['Overall']}/100 - {health_score.get('Rating', 'N/A')}</b>",
                self.styles['Normal']
            ))
            
            # Score breakdown table
            score_data = [['Category', 'Score']]
            for key, value in health_score.items():
                if key not in ['Overall', 'Rating']:
                    score_data.append([key, f"{value}/25"])
            
            score_table = Table(score_data, colWidths=[3*inch, 2*inch])
            score_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1f77b4')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 12),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            story.append(score_table)
            story.append(Spacer(1, 0.3*inch))
        
        # Add each category
        for category_name, data in self.categorized_data.items():
            story.append(PageBreak())
            story.append(Paragraph(category_name, self.styles['SectionHeader']))
            
            # Convert dataframe to table data
            table_data = [data.columns.tolist()]
            for idx, row in data.iterrows():
                table_data.append(row.tolist())
            
            # Create table
            col_widths = [2.5*inch] + [0.8*inch] * (len(data.columns) - 1)
            table = Table(table_data, colWidths=col_widths, repeatRows=1)
            
            # Style the table
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1f77b4')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (1, 0), (-1, -1), 'CENTER'),
                ('ALIGN', (0, 0), (0, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 9),
                ('FONTSIZE', (0, 1), (-1, -1), 8),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
                ('BACKGROUND', (0, 1), (-1, -1), colors.white),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey])
            ]))
            
            story.append(table)
            story.append(Spacer(1, 0.2*inch))
            
            # Add insights
            insights = self.analyzer.get_category_insights(category_name, data)
            if insights:
                story.append(Paragraph(f"<i>{insights}</i>", self.styles['Insight']))
            story.append(Spacer(1, 0.2*inch))
        
        # Build PDF
        doc.build(story)
        
        buffer.seek(0)
        return buffer.getvalue()
    
    def generate_excel_report(self) -> bytes:
        """Generate a comprehensive Excel report with multiple sheets"""
        buffer = io.BytesIO()
        
        with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
            # Summary sheet
            summary_data = []
            summary_data.append(['Qualcomm Financial Analysis Report'])
            summary_data.append(['Generated on:', datetime.now().strftime('%B %d, %Y')])
            summary_data.append([])
            summary_data.append(['Executive Summary'])
            
            insights = self.analyzer.generate_executive_summary()
            for insight in insights:
                summary_data.append([insight])
            
            summary_data.append([])
            summary_data.append(['Financial Health Score'])
            
            health_score = self.analyzer.calculate_financial_health_score()
            if health_score:
                summary_data.append(['Overall Score:', f"{health_score.get('Overall', 'N/A')}/100"])
                summary_data.append(['Rating:', health_score.get('Rating', 'N/A')])
                summary_data.append([])
                summary_data.append(['Category', 'Score'])
                for key, value in health_score.items():
                    if key not in ['Overall', 'Rating']:
                        summary_data.append([key, f"{value}/25"])
            
            summary_df = pd.DataFrame(summary_data)
            summary_df.to_excel(writer, sheet_name='Summary', index=False, header=False)
            
            # Format summary sheet
            workbook = writer.book
            summary_sheet = writer.sheets['Summary']
            
            # Title formatting
            summary_sheet['A1'].font = Font(size=18, bold=True, color='1f77b4')
            summary_sheet['A1'].alignment = Alignment(horizontal='center')
            summary_sheet.merge_cells('A1:F1')
            
            # Section headers
            for row in summary_sheet.iter_rows():
                for cell in row:
                    if cell.value in ['Executive Summary', 'Financial Health Score']:
                        cell.font = Font(size=14, bold=True, color='2c3e50')
                        cell.fill = PatternFill(start_color='e8f4f8', end_color='e8f4f8', fill_type='solid')
            
            # Adjust column widths
            summary_sheet.column_dimensions['A'].width = 50
            summary_sheet.column_dimensions['B'].width = 20
            
            # Add each category as a separate sheet
            for category_name, data in self.categorized_data.items():
                # Clean sheet name (Excel has restrictions)
                sheet_name = category_name[:31]  # Excel limit
                
                data.to_excel(writer, sheet_name=sheet_name, index=False)
                
                worksheet = writer.sheets[sheet_name]
                
                # Header formatting
                header_fill = PatternFill(start_color='1f77b4', end_color='1f77b4', fill_type='solid')
                header_font = Font(bold=True, color='FFFFFF')
                border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
                
                for cell in worksheet[1]:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    cell.border = border
                
                # Data formatting
                for row in worksheet.iter_rows(min_row=2):
                    for idx, cell in enumerate(row):
                        cell.border = border
                        cell.alignment = Alignment(horizontal='center' if idx > 0 else 'left', vertical='center')
                        
                        # Alternate row colors
                        if cell.row % 2 == 0:
                            cell.fill = PatternFill(start_color='f2f2f2', end_color='f2f2f2', fill_type='solid')
                
                # Adjust column widths
                for column in worksheet.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    
                    adjusted_width = min(max_length + 2, 50)
                    worksheet.column_dimensions[column_letter].width = adjusted_width
                
                # Add insights as a note in the first cell
                insights = self.analyzer.get_category_insights(category_name, data)
                if insights:
                    worksheet['A1'].comment = openpyxl.comments.Comment(insights, 'Financial Analyzer')
            
            # Full Data sheet
            self.full_data.to_excel(writer, sheet_name='Full Data', index=False)
            full_sheet = writer.sheets['Full Data']
            
            # Format full data sheet
            for cell in full_sheet[1]:
                cell.fill = PatternFill(start_color='1f77b4', end_color='1f77b4', fill_type='solid')
                cell.font = Font(bold=True, color='FFFFFF')
                cell.alignment = Alignment(horizontal='center', vertical='center')
        
        buffer.seek(0)
        return buffer.getvalue()
